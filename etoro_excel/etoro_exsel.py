from openpyxl import Workbook, load_workbook
from collections import Counter

print("Failu nosaukumu(Vajag ja tas butu fails no saitu etoro.com un beigā vajag but .xlsx): ")
f = input() #Lai ievadītu faila nosaukumu

wb = load_workbook(f)   #atverot failu
ws = wb['Dividends']    #atverot lapu
wl = wb['Account Summary']  #atverot lapu
max_row = ws.max_row
s = []
d = []
fin = []
dic = {}

start = wl['B7'].value  #uzzināt vērtību no šīs šūnas
finish = wl['B8'].value #uzzināt vērtību no šīs šūnas
print(start + "   -   " + finish)   

for i in range(2, max_row + 1): #cikls iet no 2 līdz maksimālajām rindiņām failā
    a = ws['B' + str(i)].value  #uzzināt vērtības no kolona
    s.append(a) #ierakstiet vērtību sarakstā

for m in range(2, max_row + 1): #cikls iet no 2 līdz maksimālajām rindiņām failā
    b = ws['C' + str(m)].value  #uzzināt vērtības no kolona
    d.append(b) #ierakstiet vērtību sarakstā

fin = list(zip(s, d)) #izveidot sarakstu, kurā apvienot divus sarakstus(saraksti s un b)

for name, devidents in fin: #cikls iet caur pāriem (name un devidents) fin sarakstā
    if name not in dic: #parbauda, ir šis atslega vārdnīcā
        dic[name] = devidents   #ja nav tad pievino šo atslegu ar vērtību
    else:
        dic[name] += devidents  #ja ir tad pievino pie šim atslegam šo vērtību

k = Counter(dic) #izveidot mainīgo ar tipa Counter
top = k.most_common(3)  #no vārdnīca dic pievieno 3 vislielakos pozicijas sarakstā top

print()
print("Top-3 pa dividentiem:")
for l in range(len(top)):   #cikls iet saraksta garumā
    print(str(l+1) + "." + " " + str(top[l]) + "$")

print(" ")
print("Visi dividenti:")
for name, dividents in dic.items(): #cikls iet caur pāriem (name un devidents) dic vārdnicā
    print(name, ':', dividents, "$")

print()
print("Kopeja summa: " + str(float('%.2f' % (sum(d)))) + "$")
print()
print()